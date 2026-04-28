import os
import json
import time
import uuid
import shutil
from typing import List, Optional, AsyncGenerator

from sqlalchemy.orm import Session
from fastapi import HTTPException

from ..entities import (
    Knowledge,
    KnowledgeGroup,
    KnowledgeChunk,
    KnowledgeSetting,
    Model,
    Provider,
)
from ..entities.factory import (
    KnowledgeGroupFactory,
    KnowledgeFactory,
    KnowledgeSettingFactory,
)
from ..schemas.knowledge import (
    KnowledgeGroupResponse,
    KnowledgeResponse,
    KnowledgeSearchResult,
    KnowledgeSettingsResponse,
    RecallTestResult,
    FileChunksResponse,
    ChunkInfo,
)
from .embedding_service import EmbeddingService, RerankService
from .chunking_service import split_text_to_chunks, split_mineru_content_list
from .vector_store_service import VectorStoreService
from ..config import get_settings
from ..utils.logger import get_logger

logger = get_logger(__name__)
settings = get_settings()

MAX_FILE_SIZE = 20 * 1024 * 1024


def _parse_image_paths(val) -> Optional[List[str]]:
    if val is None:
        return None
    if isinstance(val, list):
        return val
    if isinstance(val, str):
        try:
            parsed = json.loads(val)
            if isinstance(parsed, list):
                return parsed
        except (json.JSONDecodeError, TypeError):
            pass
    return None


def _dedup_search_results(results: List[KnowledgeSearchResult]) -> List[KnowledgeSearchResult]:
    deduped = []
    seen: dict = {}
    for r in results:
        key = r.content
        if key in seen:
            existing = seen[key]
            merged_images = list(set(
                (existing.image_paths or [])
                + ([r.content_path] if r.content_path else [])
                + ([existing.content_path] if existing.content_path else [])
                + (r.image_paths or [])
            ))
            existing.image_paths = merged_images or None
            if merged_images:
                existing.content_path = merged_images[0]
            if r.score > existing.score:
                existing.score = r.score
                existing.chunk_id = r.chunk_id
                existing.chunk_index = r.chunk_index
        else:
            all_images = list(set(
                (r.image_paths or [])
                + ([r.content_path] if r.content_path else [])
            ))
            r.image_paths = all_images or None
            if all_images:
                r.content_path = all_images[0]
            seen[key] = r
            deduped.append(r)
    return deduped

_embedding_service: Optional[EmbeddingService] = None
_rerank_service: Optional[RerankService] = None


def _get_vector_store() -> VectorStoreService:
    return VectorStoreService.get_instance()


def get_embedding_service(db: Session = None, model_id: Optional[int] = None) -> EmbeddingService:
    global _embedding_service

    if db and model_id:
        model = db.query(Model).filter(Model.id == model_id, Model.is_active == True).first()
        if model:
            provider = db.query(Provider).filter(Provider.id == model.provider_id, Provider.is_active == True).first() if model.provider_id else None
            if model.model_path:
                if _embedding_service is None:
                    _embedding_service = EmbeddingService(dimension=model.embedding_dimension or settings.EMBEDDING_DIMENSION)
                if not _embedding_service.is_loaded or _embedding_service.model_info != f"local:{model.model_path}":
                    _embedding_service.load_local_model(model.model_path)
            elif provider and provider.api_base and provider.api_key:
                if _embedding_service is None:
                    _embedding_service = EmbeddingService(dimension=model.embedding_dimension or settings.EMBEDDING_DIMENSION)
                _embedding_service.configure_api(
                    api_base=provider.api_base,
                    api_key=provider.api_key,
                    model=model.name,
                    dimension=model.embedding_dimension or 0,
                )
            return _embedding_service

    if _embedding_service is None:
        _embedding_service = EmbeddingService(dimension=settings.EMBEDDING_DIMENSION)
    return _embedding_service


def get_rerank_service(db: Session, model_id: Optional[int] = None) -> Optional[RerankService]:
    global _rerank_service

    if model_id:
        rerank_model = db.query(Model).filter(
            Model.id == model_id,
            Model.model_type == "rerank",
            Model.is_active == True,
        ).first()
    else:
        rerank_model = db.query(Model).filter(
            Model.model_type == "rerank",
            Model.is_active == True,
        ).first()
    if not rerank_model:
        return None

    if _rerank_service is None:
        _rerank_service = RerankService()

    if rerank_model.model_path:
        if not _rerank_service.is_loaded:
            _rerank_service.load_local_model(rerank_model.model_path)
    elif rerank_model.provider_id:
        provider = db.query(Provider).filter(Provider.id == rerank_model.provider_id, Provider.is_active == True).first()
        if provider and provider.api_base and provider.api_key:
            _rerank_service.configure_api(
                api_base=provider.api_base,
                api_key=provider.api_key,
                model=rerank_model.name,
            )
    return _rerank_service


def _extract_text(content_bytes: bytes, ext: str) -> str:
    ext = ext.lower()
    if ext in ("txt", "md", "csv", "json", "py", "js", "html", "css", "xml", "yaml", "yml", "log", "sql", "sh", "bat"):
        try:
            return content_bytes.decode("utf-8")
        except UnicodeDecodeError:
            return content_bytes.decode("gbk", errors="ignore")

    if ext == "pdf":
        return _extract_pdf(content_bytes)
    if ext in ("doc", "docx"):
        return _extract_docx(content_bytes, ext)
    if ext in ("xls", "xlsx"):
        return _extract_xlsx(content_bytes)

    try:
        return content_bytes.decode("utf-8")
    except UnicodeDecodeError:
        return content_bytes.decode("gbk", errors="ignore")


def _extract_pdf(content_bytes: bytes) -> str:
    try:
        import io
        from PyPDF2 import PdfReader
        reader = PdfReader(io.BytesIO(content_bytes))
        pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
        result = "\n\n".join(pages)
        if not result.strip():
            logger.warning(f"[Extract] PDF text is empty after extraction, pages={len(reader.pages)}")
        return result
    except ImportError:
        raise ImportError("PyPDF2 未安装，无法提取 PDF 文本。请执行: pip install PyPDF2")
    except Exception as e:
        logger.error(f"[Extract] PDF extraction failed: {e}", exc_info=True)
        raise RuntimeError(f"PDF 文本提取失败: {e}") from e


def _extract_docx(content_bytes: bytes, ext: str) -> str:
    try:
        import io
        if ext == "docx":
            from docx import Document
            doc = Document(io.BytesIO(content_bytes))
            return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
        else:
            return ""
    except ImportError:
        raise ImportError("python-docx 未安装，无法提取 DOCX 文本。请执行: pip install python-docx")
    except Exception as e:
        logger.error(f"[Extract] DOCX extraction failed: {e}", exc_info=True)
        raise RuntimeError(f"DOCX 文本提取失败: {e}") from e


def _extract_xlsx(content_bytes: bytes) -> str:
    try:
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content_bytes), read_only=True)
        rows = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(" | ".join(cells))
        return "\n".join(rows)
    except ImportError:
        raise ImportError("openpyxl 未安装，无法提取 XLSX 文本。请执行: pip install openpyxl")
    except Exception as e:
        logger.error(f"[Extract] XLSX extraction failed: {e}", exc_info=True)
        raise RuntimeError(f"XLSX 文本提取失败: {e}") from e


def _parse_with_mineru(file_path: str, filename: str, ext: str):
    from ..utils.mineru_parser import MineruParser

    output_dir = os.path.abspath(os.path.join(settings.UPLOAD_DIR, "mineru_output", uuid.uuid4().hex[:8]))
    os.makedirs(output_dir, exist_ok=True)
    file_path = os.path.abspath(file_path)
    logger.info(f"[Knowledge] >>> MineRU parse start: file={filename}, ext={ext}, output_dir={output_dir}")

    ext_lower = ext.lower()
    if ext_lower == "pdf":
        content_list, md_content = MineruParser.parse_pdf(
            pdf_path=file_path,
            output_dir=output_dir,
            method="auto",
            lang="ch",
        )
    elif ext_lower in ("doc", "docx", "ppt", "pptx", "xls", "xlsx"):
        try:
            content_list, md_content = MineruParser.parse_office_doc(
                doc_path=file_path,
                output_dir=output_dir,
                lang="ch",
            )
        except Exception:
            content_list, md_content = MineruParser.parse_office_doc_python(
                doc_path=file_path,
                output_dir=output_dir,
            )
    else:
        logger.warning(f"[Knowledge] MineRU skipped: unsupported ext={ext}")
        return "", None, ""

    text = md_content or ""
    logger.info(
        f"[Knowledge] <<< MineRU parse done: file={filename}, "
        f"text_len={len(text)}, content_list_items={len(content_list) if content_list else 0}"
    )
    return text, content_list, output_dir


# ── Groups ──────────────────────────────────────────────


def list_groups(user_id: int, db: Session) -> List[KnowledgeGroupResponse]:
    logger.info(f"[Knowledge] List groups: user_id={user_id}")
    groups = (
        db.query(KnowledgeGroup)
        .filter(KnowledgeGroup.user_id == user_id)
        .order_by(KnowledgeGroup.sort_order, KnowledgeGroup.created_at)
        .all()
    )
    result = []
    for g in groups:
        file_count = (
            db.query(Knowledge)
            .filter(Knowledge.group_id == g.id, Knowledge.user_id == user_id)
            .count()
        )
        resp = KnowledgeGroupResponse.model_validate(g)
        resp.file_count = file_count
        result.append(resp)
    logger.info(f"[Knowledge] Groups returned: user_id={user_id}, count={len(result)}")
    return result


def create_group(
    user_id: int, name: str, description: str = "", color: str = "#6366f1", icon: str = "Folder", db: Session = None
) -> KnowledgeGroupResponse:
    logger.info(f"[Knowledge] Create group: user_id={user_id}, name={name}")
    group = KnowledgeGroupFactory.create(
        user_id, name, description=description, color=color, icon=icon
    )
    db.add(group)
    db.commit()
    db.refresh(group)
    resp = KnowledgeGroupResponse.model_validate(group)
    resp.file_count = 0
    logger.info(f"[Knowledge] Group created: id={group.id}, name={name}")
    return resp


def update_group(
    user_id: int, group_id: int, update_data: dict, db: Session
) -> KnowledgeGroupResponse:
    logger.info(f"[Knowledge] Update group: user_id={user_id}, group_id={group_id}")
    group = (
        db.query(KnowledgeGroup)
        .filter(KnowledgeGroup.id == group_id, KnowledgeGroup.user_id == user_id)
        .first()
    )
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    for key, value in update_data.items():
        if value is not None:
            setattr(group, key, value)
    db.commit()
    db.refresh(group)
    file_count = db.query(Knowledge).filter(Knowledge.group_id == group.id).count()
    resp = KnowledgeGroupResponse.model_validate(group)
    resp.file_count = file_count
    logger.info(f"[Knowledge] Group updated: id={group_id}")
    return resp


def delete_group(user_id: int, group_id: int, db: Session):
    logger.info(f"[Knowledge] Delete group: user_id={user_id}, group_id={group_id}")
    group = (
        db.query(KnowledgeGroup)
        .filter(KnowledgeGroup.id == group_id, KnowledgeGroup.user_id == user_id)
        .first()
    )
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    files = db.query(Knowledge).filter(Knowledge.group_id == group_id).all()
    for f in files:
        f.group_id = None
    db.delete(group)
    db.commit()
    logger.info(
        f"[Knowledge] Group deleted: id={group_id}, orphaned_files={len(files)}"
    )
    return {"message": "Group deleted"}


# ── Files ───────────────────────────────────────────────


def list_files(
    user_id: int, group_id: Optional[int], db: Session
) -> List[KnowledgeResponse]:
    logger.info(f"[Knowledge] List files: user_id={user_id}, group_id={group_id}")
    q = db.query(Knowledge).filter(Knowledge.user_id == user_id)
    if group_id is not None:
        q = q.filter(Knowledge.group_id == group_id)
    files = q.order_by(Knowledge.created_at.desc()).all()
    logger.info(f"[Knowledge] Files returned: count={len(files)}")
    return [KnowledgeResponse.model_validate(k) for k in files]


def create_file_entry(
    user_id: int, name: str, db: Session, **kwargs
) -> KnowledgeResponse:
    logger.info(f"[Knowledge] Create file entry: user_id={user_id}, name={name}")
    if kwargs.get("group_id"):
        grp = (
            db.query(KnowledgeGroup)
            .filter(
                KnowledgeGroup.id == kwargs["group_id"],
                KnowledgeGroup.user_id == user_id,
            )
            .first()
        )
        if not grp:
            raise HTTPException(status_code=404, detail="Group not found")
    item = KnowledgeFactory.create(user_id, name, **kwargs)
    db.add(item)
    db.commit()
    db.refresh(item)
    logger.info(f"[Knowledge] File entry created: id={item.id}, name={name}")
    return KnowledgeResponse.model_validate(item)


def upload_file(
    user_id: int,
    filename: str,
    content_bytes: bytes,
    db: Session,
    group_id: Optional[int] = None,
) -> KnowledgeResponse:
    start_time = time.time()
    logger.info(
        f"[Knowledge] File upload start: user_id={user_id}, filename={filename}, group_id={group_id}"
    )

    file_size = len(content_bytes)
    if file_size > MAX_FILE_SIZE:
        logger.warning(
            f"[Knowledge] File too large: filename={filename}, size={file_size}"
        )
        raise HTTPException(status_code=400, detail="File size exceeds 20MB limit")

    upload_dir = settings.UPLOAD_DIR
    os.makedirs(upload_dir, exist_ok=True)

    ext = filename.rsplit(".", 1)[-1] if "." in filename else "txt"
    safe_name = f"{uuid.uuid4().hex}.{ext}"
    file_path = os.path.join(upload_dir, safe_name)
    with open(file_path, "wb") as f:
        f.write(content_bytes)
    logger.info(f"[Knowledge] File saved: path={file_path}, size={file_size}")

    _t0 = time.time()
    extract_error = None
    content_text = ""
    parse_method = ""
    mineru_output_dir = ""
    mineru_content_list = None

    if ext.lower() == "pdf":
        logger.info(f"[Knowledge] >>> Phase 1: PDF parsing (MineRU -> PyPDF2 fallback)")
        try:
            content_text, mineru_content_list, mineru_output_dir = _parse_with_mineru(
                file_path, filename, ext
            )
            parse_method = "mineru"
            logger.info(
                f"[Knowledge] <<< Phase 1 DONE (MineRU): content_length={len(content_text)}, "
                f"content_list_items={len(mineru_content_list) if mineru_content_list else 0}, "
                f"elapsed={time.time() - _t0:.2f}s"
            )
        except Exception as e:
            logger.warning(
                f"[Knowledge] !!! Phase 1 MineRU FAILED ({time.time() - _t0:.2f}s): {e}"
            )
            logger.warning(f"[Knowledge] >>> Phase 1 fallback: trying PyPDF2...")
            _t1 = time.time()
            try:
                content_text = _extract_pdf(content_bytes)
                logger.info(f"[Knowledge] <<< Phase 1 DONE (PyPDF2): content_length={len(content_text)}, elapsed={time.time() - _t1:.2f}s")
            except Exception as e2:
                extract_error = str(e2)
                logger.error(f"[Knowledge] !!! Phase 1 FAILED completely: {e2}", exc_info=True)

    elif ext.lower() in ("doc", "docx", "ppt", "pptx", "xls", "xlsx"):
        logger.info(f"[Knowledge] >>> Phase 1: Office doc parsing (MineRU)")
        try:
            content_text, mineru_content_list, mineru_output_dir = _parse_with_mineru(
                file_path, filename, ext
            )
            parse_method = "mineru"
            logger.info(f"[Knowledge] <<< Phase 1 DONE (MineRU office): content_length={len(content_text)}, elapsed={time.time() - _t0:.2f}s")
        except Exception as e:
            logger.warning(f"[Knowledge] !!! Phase 1 MineRU office FAILED ({time.time() - _t0:.2f}s): {e}")
            try:
                content_text = _extract_text(content_bytes, ext)
                logger.info(f"[Knowledge] <<< Phase 1 DONE (fallback): content_length={len(content_text)}")
            except Exception as e2:
                extract_error = str(e2)
                logger.error(f"[Knowledge] Office extraction failed: {e2}", exc_info=True)
    else:
        try:
            content_text = _extract_text(content_bytes, ext)
            logger.info(f"[Knowledge] Text extract done: content_length={len(content_text)}, elapsed={time.time() - _t0:.2f}s")
        except Exception as e:
            extract_error = str(e)
            logger.error(f"[Knowledge] Text extraction failed: file={filename}, error={e}", exc_info=True)

    item = KnowledgeFactory.create(
        user_id,
        filename,
        group_id=group_id,
        file_path=file_path,
        file_type=ext,
        file_size=file_size,
        content=content_text[:200] if len(content_text) > 200 else content_text,
    )
    if parse_method:
        item.parse_method = parse_method
    if mineru_output_dir:
        item.mineru_output_dir = mineru_output_dir
    db.add(item)
    db.commit()
    db.refresh(item)

    indexing_error = None
    if content_text or mineru_content_list:
        logger.info(
            f"[Knowledge] Starting indexing: file_id={item.id}, "
            f"content_length={len(content_text)}, has_mineru={mineru_content_list is not None}"
        )
        try:
            index_knowledge(
                item, db,
                mineru_content_list=mineru_content_list,
                mineru_output_dir=mineru_output_dir,
                full_content=content_text,
            )
            db.refresh(item)
        except Exception as e:
            logger.error(f"[Knowledge] Indexing failed: file_id={item.id}, error={e}", exc_info=True)
            indexing_error = str(e)
    elif extract_error:
        logger.info(
            f"[Knowledge] Text extraction failed, skipping indexing: file_id={item.id}"
        )

    elapsed = time.time() - start_time
    logger.info(
        f"[Knowledge] File upload complete: id={item.id}, name={filename}, "
        f"indexed={item.indexed}, chunks={item.chunk_count}, parse_method={parse_method}, elapsed={elapsed:.2f}s"
    )
    resp = KnowledgeResponse.model_validate(item)
    resp_dict = resp.model_dump()
    if extract_error:
        resp_dict["extract_error"] = extract_error
    if indexing_error:
        resp_dict["indexing_error"] = indexing_error
    return resp_dict


def update_file(
    user_id: int, file_id: int, update_data: dict, db: Session
) -> KnowledgeResponse:
    logger.info(f"[Knowledge] Update file: user_id={user_id}, file_id={file_id}")
    item = (
        db.query(Knowledge)
        .filter(Knowledge.id == file_id, Knowledge.user_id == user_id)
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="File not found")
    for key, value in update_data.items():
        if value is not None:
            setattr(item, key, value)
    db.commit()
    db.refresh(item)
    logger.info(f"[Knowledge] File updated: id={file_id}")
    return KnowledgeResponse.model_validate(item)


def delete_file(user_id: int, file_id: int, db: Session):
    logger.info(f"[Knowledge] Delete file: user_id={user_id}, file_id={file_id}")
    item = (
        db.query(Knowledge)
        .filter(Knowledge.id == file_id, Knowledge.user_id == user_id)
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="File not found")
    chunk_count = (
        db.query(KnowledgeChunk).filter(KnowledgeChunk.knowledge_id == item.id).count()
    )
    db.query(KnowledgeChunk).filter(KnowledgeChunk.knowledge_id == item.id).delete()

    try:
        vs = _get_vector_store()
        vs.delete_by_knowledge_id(item.id)
    except Exception as e:
        logger.warning(f"[Knowledge] VectorStore delete failed: {e}")

    if item.file_path and os.path.exists(item.file_path):
        try:
            os.remove(item.file_path)
        except Exception as e:
            logger.warning(f"[Knowledge] Failed to delete file: {item.file_path}, {e}")

    if item.mineru_output_dir and os.path.isdir(item.mineru_output_dir):
        try:
            shutil.rmtree(item.mineru_output_dir)
        except Exception as e:
            logger.warning(f"[Knowledge] Failed to delete mineru output dir: {item.mineru_output_dir}, {e}")

    db.delete(item)
    db.commit()
    logger.info(
        f"[Knowledge] File deleted: id={file_id}, removed_chunks={chunk_count}"
    )
    return {"message": "File deleted"}


def reindex_file(user_id: int, file_id: int, db: Session):
    start_time = time.time()
    logger.info(f"[Knowledge] Reindex file: user_id={user_id}, file_id={file_id}")
    item = (
        db.query(Knowledge)
        .filter(Knowledge.id == file_id, Knowledge.user_id == user_id)
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="File not found")
    db.query(KnowledgeChunk).filter(KnowledgeChunk.knowledge_id == item.id).delete()
    db.commit()

    try:
        vs = _get_vector_store()
        vs.delete_by_knowledge_id(item.id)
    except Exception as e:
        logger.warning(f"[Knowledge] VectorStore delete failed: {e}")

    if not item.file_path or not os.path.exists(item.file_path):
        raise HTTPException(status_code=400, detail="原文件不存在，无法重新索引")

    mineru_content_list = None
    mineru_output_dir = item.mineru_output_dir
    content_text = ""

    ext_lower = (item.file_type or "").lower()
    if ext_lower in ("pdf", "doc", "docx", "ppt", "pptx", "xls", "xlsx"):
        try:
            content_text, mineru_content_list, mineru_output_dir = _parse_with_mineru(
                item.file_path, item.name, item.file_type
            )
            item.parse_method = "mineru"
            item.mineru_output_dir = mineru_output_dir
            logger.info(f"[Knowledge] Re-parsed with MineRU: content_length={len(content_text)}")
        except Exception as e:
            logger.warning(f"[Knowledge] MineRU re-parse failed: {e}")

    if not content_text:
        try:
            with open(item.file_path, "rb") as f:
                content_bytes = f.read()
            content_text = _extract_text(content_bytes, item.file_type)
            logger.info(f"[Knowledge] Re-extracted text: length={len(content_text)}")
        except Exception as e:
            logger.error(f"[Knowledge] Re-extract failed: file_id={file_id}, error={e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"文本提取失败：{e}")

    item.content = content_text[:200] if len(content_text) > 200 else content_text
    db.commit()

    if content_text or mineru_content_list:
        try:
            index_knowledge(
                item, db,
                mineru_content_list=mineru_content_list,
                mineru_output_dir=mineru_output_dir,
                full_content=content_text,
            )
            db.refresh(item)
        except Exception as e:
            logger.error(f"[Knowledge] Reindex failed: file_id={file_id}, error={e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"索引失败：{e}")
        elapsed = time.time() - start_time
        logger.info(
            f"[Knowledge] Reindex complete: file_id={file_id}, chunks={item.chunk_count}, elapsed={elapsed:.2f}s"
        )
        return {"message": "Reindexed", "chunks": item.chunk_count}
    logger.info(
        f"[Knowledge] Reindex skipped (no content): file_id={file_id}"
    )
    return {"message": "No content to index"}


def get_file_chunks(user_id: int, file_id: int, db: Session) -> FileChunksResponse:
    logger.info(f"[Knowledge] Get file chunks: user_id={user_id}, file_id={file_id}")
    item = (
        db.query(Knowledge)
        .filter(Knowledge.id == file_id, Knowledge.user_id == user_id)
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="File not found")

    try:
        vs = _get_vector_store()
        raw_chunks = vs.get_chunks_by_knowledge_id(item.id)
    except Exception as e:
        logger.warning(f"[Knowledge] ChromaDB get_chunks failed: {e}")
        raw_chunks = []

    chunk_infos = [
        ChunkInfo(
            chunk_index=c["chunk_index"],
            content=c["content"],
            char_count=c["char_count"],
        )
        for c in raw_chunks
    ]

    return FileChunksResponse(
        file_id=item.id,
        file_name=item.name,
        total_chunks=len(chunk_infos),
        chunks=chunk_infos,
    )


# ── Search ──────────────────────────────────────────────


def search_knowledge(
    user_id: int, query: str, group_id: Optional[int], top_k: int, db: Session
) -> List[KnowledgeSearchResult]:
    start_time = time.time()
    logger.info(
        f"[Knowledge] Search: user_id={user_id}, query_length={len(query)}, group_id={group_id}, top_k={top_k}"
    )

    ks = get_or_create_settings(user_id, db, group_id=group_id)
    emb_svc = get_embedding_service(db=db, model_id=ks.embedding_model_id)
    try:
        query_vec = emb_svc.embed_query(query)
    except Exception as e:
        logger.error(f"[Knowledge] Embedding failed in search: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Embedding 生成失败：{e}")

    try:
        vs = _get_vector_store()
        results = vs.search(
            query_embedding=query_vec,
            user_id=user_id,
            group_id=group_id,
            top_k=top_k * 3,
        )
    except Exception as e:
        logger.error(f"[Knowledge] VectorStore search failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"搜索失败：{e}")

    if not results:
        logger.info("[Knowledge] Search: no results found")
        return []

    knowledge_cache = {}
    scored = []
    for r in results:
        kid = r["metadata"].get("knowledge_id", 0)
        if kid not in knowledge_cache:
            knowledge_cache[kid] = (
                db.query(Knowledge).filter(Knowledge.id == kid).first()
            )
        kn = knowledge_cache.get(kid)
        scored.append(
            KnowledgeSearchResult(
                chunk_id=r["id"],
                knowledge_id=kid,
                knowledge_name=kn.name if kn else "",
                content=r["content"],
                score=r["score"],
                chunk_index=r["metadata"].get("chunk_index", 0),
                chunk_type=r["metadata"].get("chunk_type", "text"),
                page_idx=r["metadata"].get("page_idx"),
                content_path=r["metadata"].get("content_path"),
                image_path=r["metadata"].get("image_path"),
                image_paths=_parse_image_paths(r["metadata"].get("image_paths")),
            )
        )

    scored = _dedup_search_results(scored)

    if ks.enable_rerank:
        rerank_svc = get_rerank_service(db, model_id=ks.rerank_model_id)
        if rerank_svc and len(scored) > 0:
            candidate_count = min(len(scored), top_k * 3)
            candidates = scored[:candidate_count]
            rerank_results = rerank_svc.rerank(
                query, [c.content for c in candidates], top_k=top_k
            )
            reranked = []
            for rr in rerank_results:
                idx = rr.get("index", 0)
                if idx < len(candidates):
                    item = candidates[idx]
                    item.score = rr.get("score", item.score)
                    reranked.append(item)
            scored = reranked

    elapsed = time.time() - start_time
    logger.info(
        f"[Knowledge] Search complete: results={len(scored[:top_k])}, top_score={scored[0].score if scored else 0:.4f}, elapsed={elapsed:.2f}s"
    )
    return scored[:top_k]


# ── Settings ───────────────────────────────────────────


def get_or_create_settings(user_id: int, db: Session, group_id: Optional[int] = None) -> KnowledgeSetting:
    q = db.query(KnowledgeSetting).filter(KnowledgeSetting.user_id == user_id)
    if group_id is not None:
        q = q.filter(KnowledgeSetting.group_id == group_id)
    else:
        q = q.filter(KnowledgeSetting.group_id.is_(None))
    ks = q.first()
    if not ks:
        ks = KnowledgeSettingFactory.create(user_id, group_id=group_id)
        db.add(ks)
        db.commit()
        db.refresh(ks)
    return ks


def get_settings_response(user_id: int, db: Session, group_id: Optional[int] = None) -> KnowledgeSettingsResponse:
    ks = get_or_create_settings(user_id, db, group_id=group_id)
    return KnowledgeSettingsResponse.model_validate(ks)


def update_settings(user_id: int, update_data: dict, db: Session) -> KnowledgeSettingsResponse:
    logger.info(f"[Knowledge] Update settings: user_id={user_id}")
    group_id = update_data.pop("group_id", None)
    ks = get_or_create_settings(user_id, db, group_id=group_id)
    for key, value in update_data.items():
        if value is not None:
            setattr(ks, key, value)
    db.commit()
    db.refresh(ks)
    logger.info(f"[Knowledge] Settings updated: user_id={user_id}, group_id={group_id}")
    return KnowledgeSettingsResponse.model_validate(ks)


def list_embedding_models(db: Session):
    models = db.query(Model).filter(
        Model.is_active == True,
        Model.model_type.in_(["embedding", "rerank"]),
    ).order_by(Model.model_type, Model.name).all()
    return [
        {
            "id": m.id,
            "name": m.name,
            "display_name": m.display_name or m.name,
            "model_type": m.model_type,
            "model_path": m.model_path or "",
            "embedding_dimension": m.embedding_dimension or 0,
        }
        for m in models
    ]


# ── Recall Test ────────────────────────────────────────


def recall_test(
    user_id: int, query: str, top_k: Optional[int], db: Session, group_id: Optional[int] = None
) -> List[dict]:
    start_time = time.time()
    logger.info(
        f"[Knowledge] Recall test: user_id={user_id}, query_length={len(query)}, top_k={top_k}, group_id={group_id}"
    )

    ks = get_or_create_settings(user_id, db, group_id=group_id)
    top_k = top_k or ks.retrieval_top_k or 5
    emb_svc = get_embedding_service(db=db, model_id=ks.embedding_model_id)
    try:
        query_vec = emb_svc.embed_query(query)
    except Exception as e:
        logger.error(f"[Knowledge] Embedding failed in recall test: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Embedding 生成失败：{e}")

    try:
        vs = _get_vector_store()
        results = vs.search(
            query_embedding=query_vec,
            user_id=user_id,
            group_id=group_id,
            top_k=top_k,
        )
    except Exception as e:
        logger.error(f"[Knowledge] VectorStore search failed in recall test: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"向量检索失败：{e}")

    knowledge_cache = {}
    scored = []
    for r in results:
        kid = r["metadata"].get("knowledge_id", 0)
        if kid not in knowledge_cache:
            knowledge_cache[kid] = (
                db.query(Knowledge).filter(Knowledge.id == kid).first()
            )
        kn = knowledge_cache.get(kid)
        scored.append(
            {
                "chunk_id": r["id"],
                "knowledge_id": kid,
                "knowledge_name": kn.name if kn else "",
                "content": r["content"],
                "score": r["score"],
                "chunk_index": r["metadata"].get("chunk_index", 0),
                "retrieval_method": ks.retrieval_method,
            }
        )

    if ks.enable_rerank:
        rerank_svc = get_rerank_service(db, model_id=ks.rerank_model_id)
        if rerank_svc and len(scored) > 0:
            candidate_count = min(len(scored), top_k * 3)
            candidates = scored[:candidate_count]
            rerank_results = rerank_svc.rerank(
                query, [c["content"] for c in candidates], top_k=top_k
            )
            reranked = []
            for rr in rerank_results:
                idx = rr.get("index", 0)
                if idx < len(candidates):
                    item = candidates[idx]
                    item["score"] = rr.get("score", item["score"])
                    reranked.append(item)
            scored = reranked

    elapsed = time.time() - start_time
    logger.info(
        f"[Knowledge] Recall test complete: results={len(scored[:top_k])}, top_score={scored[0]['score'] if scored else 0:.4f}, elapsed={elapsed:.2f}s"
    )
    return scored[:top_k]


# ── Indexing ────────────────────────────────────────────


def index_knowledge(
    item: Knowledge,
    db: Session,
    mineru_content_list: Optional[List[dict]] = None,
    mineru_output_dir: str = "",
    full_content: str = "",
):
    start_time = time.time()
    logger.info(
        f"[Knowledge] Indexing: file_id={item.id}, content_length={len(full_content)}, "
        f"has_mineru={mineru_content_list is not None}"
    )

    ks = None
    if item.group_id:
        ks = db.query(KnowledgeSetting).filter(
            KnowledgeSetting.user_id == item.user_id,
            KnowledgeSetting.group_id == item.group_id,
        ).first()
    if not ks:
        ks = db.query(KnowledgeSetting).filter(
            KnowledgeSetting.user_id == item.user_id,
            KnowledgeSetting.group_id.is_(None),
        ).first()
    if not ks:
        logger.warning(f"[Knowledge] No settings found for user_id={item.user_id}, using defaults")
    chunk_method = ks.chunk_method if ks else "auto"
    chunk_size = ks.chunk_size if ks else settings.CHUNK_SIZE
    chunk_overlap = ks.chunk_overlap if ks else settings.CHUNK_OVERLAP
    embedding_model_id = ks.embedding_model_id if ks else None

    chunks_info = []
    logger.info(f"[Knowledge] >>> Phase 2: Chunking text (method={chunk_method}, size={chunk_size}, overlap={chunk_overlap})")

    if mineru_content_list:
        chunks_info = split_mineru_content_list(
            content_list=mineru_content_list,
            output_dir=mineru_output_dir or item.mineru_output_dir or "",
            document_name=item.name,
            max_chunk_size=chunk_size,
            min_chunk_size=max(chunk_size // 10, 50),
            overlap_size=chunk_overlap,
        )
        logger.info(f"[Knowledge] MineRU chunks: file_id={item.id}, count={len(chunks_info)}")

    if not chunks_info:
        chunks_info = split_text_to_chunks(
            full_content,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            method=chunk_method,
        )

    if not chunks_info:
        logger.warning(f"[Knowledge] No chunks generated: file_id={item.id}, content_length={len(full_content)}, method={chunk_method}, chunk_size={chunk_size}")
        return

    logger.info(f"[Knowledge] <<< Phase 2 DONE: Chunks generated: file_id={item.id}, count={len(chunks_info)}, elapsed={time.time() - start_time:.2f}s")
    _t_embed = time.time()
    emb_svc = get_embedding_service(db=db, model_id=embedding_model_id)
    logger.info(f"[Knowledge] >>> Phase 3: Generating embeddings: chunks={len(chunks_info)}, model_id={embedding_model_id}, emb_svc_type={type(emb_svc).__name__}, is_loaded={emb_svc.is_loaded}")

    texts_to_embed = []
    for c in chunks_info:
        texts_to_embed.append(c.get("content", ""))

    embeddings = emb_svc.embed_texts(texts_to_embed)
    logger.info(f"[Knowledge] <<< Phase 3 DONE: embeddings count={len(embeddings)}, elapsed={time.time() - _t_embed:.2f}s")

    _t_store = time.time()
    logger.info(f"[Knowledge] >>> Phase 4: Saving chunks to VectorStore + DB")
    chunk_ids = []
    documents = []
    embedding_list = []
    metadatas = []

    for i, (info, vec) in enumerate(zip(chunks_info, embeddings)):
        chunk = KnowledgeFactory.create_chunk(
            knowledge_id=item.id,
            chunk_index=info.get("chunk_index", i),
            group_id=item.group_id,
        )
        chunk.chunk_type = info.get("chunk_type", "text")
        chunk.page_idx = info.get("page_idx")
        chunk.content_path = info.get("content_path", "")
        chunk.image_path = info.get("image_path", "")
        db.add(chunk)

        chunk_id = f"{item.id}_{info.get('chunk_index', i)}"
        meta: dict = {
            "user_id": item.user_id,
            "knowledge_id": item.id,
            "chunk_index": info.get("chunk_index", i),
            "chunk_type": info.get("chunk_type", "text"),
        }
        if item.group_id is not None:
            meta["group_id"] = item.group_id
        if info.get("page_idx") is not None:
            meta["page_idx"] = info["page_idx"]
        if info.get("content_path"):
            meta["content_path"] = info["content_path"]
        if info.get("image_path"):
            meta["image_path"] = info["image_path"]
        img_paths_list = info.get("image_paths", [])
        if img_paths_list:
            meta["image_paths"] = json.dumps(img_paths_list, ensure_ascii=False)
        extra_meta = info.get("metadata", {})
        if extra_meta:
            meta.update(extra_meta)

        chunk_ids.append(chunk_id)
        documents.append(info.get("content", ""))
        embedding_list.append(vec)
        metadatas.append(meta)

    try:
        vs = _get_vector_store()
        vs.add_chunks(
            ids=chunk_ids,
            documents=documents,
            embeddings=embedding_list,
            metadatas=metadatas,
        )
    except Exception as e:
        logger.error(f"[Knowledge] !!! Phase 4 FAILED: VectorStore add failed: {e}")
    logger.info(f"[Knowledge] <<< Phase 4 DONE: chunks={len(chunk_ids)}, elapsed={time.time() - _t_store:.2f}s")

    item.chunk_count = len(chunks_info)
    item.indexed = True
    if len(item.content) > 200:
        item.content = item.content[:200]
    db.commit()

    elapsed = time.time() - start_time
    logger.info(
        f"[Knowledge] Indexing complete: file_id={item.id}, chunks={len(chunks_info)}, elapsed={elapsed:.2f}s"
    )


def search_for_agent(user_id: int, query: str, group_id: Optional[int], top_k: int, db: Session) -> List[dict]:
    results = search_knowledge(user_id, query, group_id, top_k, db)
    return [
        {
            "content": r.content,
            "source": r.knowledge_name,
            "score": round(r.score, 4),
            "chunk_index": r.chunk_index,
            "page_idx": r.page_idx,
            "content_path": r.content_path,
            "image_paths": r.image_paths,
        }
        for r in results
    ]


def rag_search_with_chunks(
    user_id: int, query: str, group_id: Optional[int], top_k: int, db: Session
) -> List[KnowledgeSearchResult]:
    return search_knowledge(user_id, query, group_id, top_k, db)


async def rag_answer_stream(
    user_id: int,
    query: str,
    search_results: List[KnowledgeSearchResult],
    db: Session,
    model_id: Optional[int] = None,
) -> AsyncGenerator[str, None]:
    from .model_service import resolve_model
    from openai import AsyncOpenAI

    model, provider = resolve_model(db, "chat", model_id)

    ref_parts = []
    for i, r in enumerate(search_results):
        ref_item = f"[参考资料{i+1}] 来源: {r.knowledge_name}"
        if r.page_idx is not None:
            ref_item += f", 第{r.page_idx + 1}页"
        ref_item += f"\n{r.content}"

        img_refs = r.image_paths or ([r.content_path] if r.content_path and r.content_path.startswith("/") else [])
        if img_refs:
            ref_item += "\n" + "\n".join(f"[关联图片: {p}]" for p in img_refs if p)

        ref_parts.append(ref_item)

    references = "\n\n".join(ref_parts)

    system_prompt = (
        "你是一个专业的知识库问答助手。请根据提供的参考资料回答用户的问题。\n"
        "回答要求：\n"
        "1. 基于参考资料内容回答，如果参考资料不足以回答，请明确说明\n"
        "2. 引用资料时使用 [参考资料X] 格式标注来源\n"
        "3. 回答要准确、完整、有条理\n"
        "4. 如果参考资料中包含[关联图片]，你必须在回答中用 Markdown 图片语法展示该图片，格式为：![图片描述](图片URL)\n"
        "5. 使用中文回答\n"
        "6. 必须展示所有关联图片，不要省略"
    )

    user_msg = f"用户问题：{query}\n\n参考资料：\n{references}"

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_msg},
    ]

    client = AsyncOpenAI(base_url=provider.api_base, api_key=provider.api_key)

    try:
        stream = await client.chat.completions.create(
            model=model.name,
            messages=messages,
            max_tokens=model.max_tokens if model.max_tokens else 4096,
            temperature=float(model.temperature) if model.temperature else 0.7,
            stream=True,
        )
        async for chunk in stream:
            delta = chunk.choices[0].delta
            if delta.content:
                yield delta.content
    except Exception as e:
        logger.error(f"[Knowledge RAG] LLM stream failed: {e}")
        yield f"\n\n[错误] AI 服务调用失败: {str(e)}"
